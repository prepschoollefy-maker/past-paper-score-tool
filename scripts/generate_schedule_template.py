"""受験計画データ用Excelテンプレート生成スクリプト（簡略版）"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "受験スケジュール"

# ── 列定義 (必要最小限) ──
COLUMNS = [
    # (ヘッダー, 幅, 必須, 説明, サンプル値)
    ("学校名",        20, True,  "正式名称",              "開成中学校"),
    ("年度",          8,  True,  "入試年度(西暦)",        2026),
    ("回次",          12, True,  "第1回, A日程 等",       "第1回"),
    ("試験日",        14, True,  "YYYY-MM-DD",            "2026-02-01"),
    ("午前午後",      10, True,  "午前 or 午後",          "午前"),
    ("性別区分",      10, True,  "男子校/女子校/共学",    "男子校"),
    ("集合時間",      10, True,  "HH:MM",                 "8:00"),
    ("四谷80偏差値",  12, True,  "整数",                  66),
    ("合格発表日時",  20, True,  "YYYY-MM-DD HH:MM",     "2026-02-02 12:00"),
    ("入学金",        12, True,  "円",                    300000),
    # ── 以下は任意 ──
    ("終了予定",      10, False, "HH:MM",                 "12:30"),
    ("科目",          14, False, "算国理社/算国 等",      "算国理社"),
    ("入学手続き締切",20, False, "YYYY-MM-DD HH:MM",     "2026-02-03 15:00"),
    ("延納可否",      10, False, "可/否",                 "否"),
    ("延納金",        10, False, "円(延納可の場合)",      ""),
]

# ── スタイル ──
fill_req = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
fill_opt = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
font_req = Font(bold=True, color="FFFFFF", size=10)
font_opt = Font(bold=True, color="134E4A", size=10)
desc_fill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")
desc_font = Font(color="6B7280", size=9, italic=True)
sample_font = Font(color="9CA3AF", size=10)
border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# ── ヘッダー行 ──
for i, (name, width, req, desc, sample) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=i, value=name)
    cell.fill = fill_req if req else fill_opt
    cell.font = font_req if req else font_opt
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = width

# ── 説明行 ──
for i, (name, width, req, desc, sample) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=2, column=i, value=f"{'【必須】' if req else '【任意】'} {desc}")
    cell.fill = desc_fill
    cell.font = desc_font
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[2].height = 30

# ── サンプルデータ ──
SAMPLES = [
    ["開成中学校",           2026, "第1回", "2026-02-01", "午前", "男子校", "8:00",  66, "2026-02-02 12:00", 300000, "12:30", "算国理社", "2026-02-03 15:00", "否", ""],
    ["渋谷教育学園渋谷中学校", 2026, "第1回", "2026-02-01", "午前", "共学",   "8:30",  58, "2026-02-01 22:00", 330000, "12:00", "算国理社", "2026-02-03 12:00", "可",  100000],
    ["栄東中学校",           2026, "A日程", "2026-01-10", "午前", "共学",   "8:30",  55, "2026-01-10 18:00", 250000, "12:00", "算国理社", "2026-01-15 23:59", "可",  50000],
]

for r, row_data in enumerate(SAMPLES, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = sample_font
        cell.border = border

# ── 入力規則 ──
dv_period = DataValidation(type="list", formula1='"午前,午後"', allow_blank=False)
dv_period.error = "午前 または 午後"
ws.add_data_validation(dv_period)
dv_period.add("E3:E500")

dv_gender = DataValidation(type="list", formula1='"男子校,女子校,共学"', allow_blank=False)
ws.add_data_validation(dv_gender)
dv_gender.add("F3:F500")

# 科目 (よく使うパターンをドロップダウン + 手入力も可)
dv_subj = DataValidation(type="list", formula1='"算国理社,算国,算英,算国英,算国理,算国社,算,適性検査"', allow_blank=True)
dv_subj.prompt = "選択または直接入力"
dv_subj.promptTitle = "科目"
dv_subj.showInputMessage = True
ws.add_data_validation(dv_subj)
dv_subj.add("L3:L500")

# 延納可否
dv_defer = DataValidation(type="list", formula1='"可,否"', allow_blank=True)
ws.add_data_validation(dv_defer)
dv_defer.add("N3:N500")

# ── 固定・フィルター ──
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# ── 保存 ──
path = "public/template_schedule_data.xlsx"
wb.save(path)
print(f"テンプレート生成完了: {path}")
print(f"列数: {len(COLUMNS)} (必須{sum(1 for c in COLUMNS if c[2])}列 + 任意{sum(1 for c in COLUMNS if not c[2])}列)")
